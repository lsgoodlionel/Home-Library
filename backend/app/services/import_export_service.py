from __future__ import annotations

import csv
import io
import json
from collections import Counter
from collections.abc import Iterable
from typing import Any

from pydantic import ValidationError
from sqlalchemy.orm import Session, joinedload, selectinload

from app.core.errors import ApiError
from app.models import Book, BookTag, Category, Location
from app.schemas.book import BookCreate
from app.schemas.import_export import (
    ImportConfirmResponse,
    ImportPreviewResponse,
    ImportPreviewRow,
    ImportRowIssue,
)
from app.services import book_service

BOOK_FIELDS = [
    "title",
    "subtitle",
    "author",
    "translator",
    "publisher",
    "publish_year",
    "isbn",
    "original_isbn",
    "language",
    "pages",
    "price_cents",
    "binding",
    "series",
    "cover_url",
    "summary",
    "author_intro",
    "category_id",
    "category_code",
    "category_name",
    "location_id",
    "location_full_path",
    "status",
    "read_status",
    "rating",
    "is_favorite",
    "tag_names",
    "note",
]

EXPORT_FIELDS = [
    "title",
    "subtitle",
    "author",
    "translator",
    "publisher",
    "publish_year",
    "isbn",
    "original_isbn",
    "language",
    "pages",
    "price_cents",
    "binding",
    "series",
    "cover_url",
    "summary",
    "author_intro",
    "category_id",
    "category_code",
    "category_name",
    "location_id",
    "location_full_path",
    "status",
    "read_status",
    "rating",
    "is_favorite",
    "tag_names",
    "note",
]

DEFAULT_FIELD_MAPPING = {
    "书名": "title",
    "标题": "title",
    "副标题": "subtitle",
    "作者": "author",
    "译者": "translator",
    "出版社": "publisher",
    "出版年份": "publish_year",
    "ISBN": "isbn",
    "isbn": "isbn",
    "原始ISBN": "original_isbn",
    "语言": "language",
    "页数": "pages",
    "价格分": "price_cents",
    "定价分": "price_cents",
    "装帧": "binding",
    "丛书": "series",
    "封面": "cover_url",
    "简介": "summary",
    "作者简介": "author_intro",
    "分类ID": "category_id",
    "分类号": "category_code",
    "分类名": "category_name",
    "位置ID": "location_id",
    "位置": "location_full_path",
    "状态": "status",
    "阅读状态": "read_status",
    "评分": "rating",
    "重点收藏": "is_favorite",
    "标签": "tag_names",
    "备注": "note",
}


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _clean_isbn(value: Any) -> str | None:
    text = _clean_text(value)
    if not text:
        return None
    cleaned = "".join(char for char in text if char.isdigit() or char.upper() == "X")
    return cleaned or None


def _looks_like_isbn(value: str) -> bool:
    return (len(value) == 10 and all(c.isdigit() or c == "X" for c in value.upper())) or (
        len(value) == 13 and value.isdigit()
    )


def _to_int(value: Any) -> int | None:
    text = _clean_text(value)
    if text is None:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = (_clean_text(value) or "").lower()
    return text in {"1", "true", "yes", "y", "是", "重点", "收藏"}


def _split_tags(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    text = _clean_text(value)
    if not text:
        return []
    for separator in [";", "；", ",", "，"]:
        text = text.replace(separator, "|")
    return [item.strip() for item in text.split("|") if item.strip()]


def _decode_content(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ApiError("VALIDATION_ERROR", "文件编码无法识别，请使用 UTF-8 或 GB18030", status_code=422)


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = _decode_content(content)
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ApiError("VALIDATION_ERROR", "CSV 文件缺少表头", status_code=422)
    return [dict(row) for row in reader]


def _parse_json(content: bytes) -> list[dict[str, Any]]:
    try:
        payload = json.loads(_decode_content(content))
    except json.JSONDecodeError as exc:
        raise ApiError("VALIDATION_ERROR", f"JSON 格式错误：{exc.msg}", status_code=422) from exc

    if isinstance(payload, dict):
        payload = payload.get("items", payload.get("books", payload.get("rows")))
    if not isinstance(payload, list):
        raise ApiError("VALIDATION_ERROR", "JSON 导入文件必须是数组，或包含 items/books/rows 数组", status_code=422)
    if not all(isinstance(item, dict) for item in payload):
        raise ApiError("VALIDATION_ERROR", "JSON 导入数组中的每一项都必须是对象", status_code=422)
    return payload


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ApiError("VALIDATION_ERROR", "Excel 导入需要安装 openpyxl 依赖", status_code=422) from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_clean_text(value) or "" for value in rows[0]]
    if not any(headers):
        raise ApiError("VALIDATION_ERROR", "Excel 文件缺少表头", status_code=422)
    parsed: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(_clean_text(value) for value in row):
            continue
        parsed.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
    return parsed


def parse_import_file(content: bytes, file_format: str) -> list[dict[str, Any]]:
    normalized = file_format.lower()
    if normalized == "csv":
        return _parse_csv(content)
    if normalized == "json":
        return _parse_json(content)
    if normalized in {"xlsx", "excel"}:
        return _parse_xlsx(content)
    raise ApiError("VALIDATION_ERROR", "format 仅支持 csv、json、xlsx", status_code=422)


def _apply_field_mapping(row: dict[str, Any], field_mapping: dict[str, str] | None) -> dict[str, Any]:
    mapped: dict[str, Any] = {}
    mapping = {**DEFAULT_FIELD_MAPPING, **(field_mapping or {})}
    for source, value in row.items():
        target = mapping.get(source, source)
        if target in BOOK_FIELDS:
            mapped[target] = value

    if field_mapping:
        for target, source in field_mapping.items():
            if target in BOOK_FIELDS and source in row:
                mapped[target] = row[source]
    return mapped


def _lookup_category_id(db: Session, data: dict[str, Any], issues: list[ImportRowIssue], row_number: int) -> int | None:
    category_id = _to_int(data.get("category_id"))
    if category_id is not None:
        if db.get(Category, category_id) is None:
            issues.append(
                ImportRowIssue(row_number=row_number, field="category_id", code="CATEGORY_NOT_FOUND", message="分类不存在")
            )
        return category_id

    code = _clean_text(data.get("category_code"))
    name = _clean_text(data.get("category_name"))
    category = None
    if code:
        category = db.query(Category).filter(Category.code == code).first()
    elif name:
        category = db.query(Category).filter(Category.name == name).first()
    if (code or name) and category is None:
        issues.append(
            ImportRowIssue(row_number=row_number, field="category_id", code="CATEGORY_NOT_FOUND", message="分类不存在")
        )
    return category.id if category else None


def _lookup_location_id(db: Session, data: dict[str, Any], issues: list[ImportRowIssue], row_number: int) -> int | None:
    location_id = _to_int(data.get("location_id"))
    if location_id is not None:
        if db.get(Location, location_id) is None:
            issues.append(
                ImportRowIssue(row_number=row_number, field="location_id", code="LOCATION_NOT_FOUND", message="位置不存在")
            )
        return location_id

    full_path = _clean_text(data.get("location_full_path"))
    if not full_path:
        return None
    location = db.query(Location).filter(Location.full_path == full_path).first()
    if location is None:
        issues.append(
            ImportRowIssue(row_number=row_number, field="location_id", code="LOCATION_NOT_FOUND", message="位置不存在")
        )
        return None
    return location.id


def _normalize_row(
    db: Session,
    row: dict[str, Any],
    row_number: int,
    field_mapping: dict[str, str] | None,
) -> tuple[dict[str, Any], list[ImportRowIssue], list[ImportRowIssue]]:
    mapped = _apply_field_mapping(row, field_mapping)
    errors: list[ImportRowIssue] = []
    warnings: list[ImportRowIssue] = []

    title = _clean_text(mapped.get("title"))
    if not title:
        errors.append(ImportRowIssue(row_number=row_number, field="title", code="MISSING_TITLE", message="缺少书名"))

    isbn = _clean_isbn(mapped.get("isbn"))
    if isbn and not _looks_like_isbn(isbn):
        warnings.append(
            ImportRowIssue(
                row_number=row_number,
                field="isbn",
                code="ISBN_FORMAT_WARNING",
                message="ISBN 格式可能不正确，应为 10 位或 13 位",
                severity="warning",
            )
        )

    data = {
        "title": title or "",
        "subtitle": _clean_text(mapped.get("subtitle")),
        "author": _clean_text(mapped.get("author")),
        "translator": _clean_text(mapped.get("translator")),
        "publisher": _clean_text(mapped.get("publisher")),
        "publish_year": _to_int(mapped.get("publish_year")),
        "isbn": isbn,
        "original_isbn": _clean_text(mapped.get("original_isbn")) or _clean_text(mapped.get("isbn")),
        "language": _clean_text(mapped.get("language")),
        "pages": _to_int(mapped.get("pages")),
        "price_cents": _to_int(mapped.get("price_cents")),
        "binding": _clean_text(mapped.get("binding")),
        "series": _clean_text(mapped.get("series")),
        "cover_url": _clean_text(mapped.get("cover_url")),
        "summary": _clean_text(mapped.get("summary")),
        "author_intro": _clean_text(mapped.get("author_intro")),
        "category_id": _lookup_category_id(db, mapped, errors, row_number),
        "location_id": _lookup_location_id(db, mapped, errors, row_number),
        "status": _clean_text(mapped.get("status")) or "available",
        "read_status": _clean_text(mapped.get("read_status")) or "unread",
        "rating": _to_int(mapped.get("rating")),
        "is_favorite": _to_bool(mapped.get("is_favorite")),
        "tag_names": _split_tags(mapped.get("tag_names")),
        "note": _clean_text(mapped.get("note")),
        "source": "import",
    }

    try:
        BookCreate(**data)
    except ValidationError as exc:
        for error in exc.errors():
            field = str(error["loc"][0]) if error.get("loc") else None
            errors.append(
                ImportRowIssue(
                    row_number=row_number,
                    field=field,
                    code="VALIDATION_ERROR",
                    message=str(error.get("msg", "字段校验失败")),
                )
            )
    return data, errors, warnings


def _add_duplicate_isbn_issues(db: Session, rows: list[ImportPreviewRow]) -> None:
    isbns = [row.data.get("isbn") for row in rows if row.data.get("isbn")]
    counts = Counter(isbns)
    existing = {
        isbn
        for (isbn,) in db.query(Book.isbn).filter(Book.isbn.in_(isbns)).all()
        if isbn
    }
    for row in rows:
        isbn = row.data.get("isbn")
        if not isbn:
            continue
        if counts[isbn] > 1 or isbn in existing:
            row.errors.append(
                ImportRowIssue(
                    row_number=row.row_number,
                    field="isbn",
                    code="DUPLICATE_ISBN",
                    message="重复 ISBN",
                )
            )


def preview_import(
    db: Session,
    content: bytes,
    file_format: str,
    field_mapping: dict[str, str] | None = None,
) -> ImportPreviewResponse:
    raw_rows = parse_import_file(content, file_format)
    rows: list[ImportPreviewRow] = []
    for index, raw_row in enumerate(raw_rows, start=2):
        data, errors, warnings = _normalize_row(db, raw_row, index, field_mapping)
        rows.append(ImportPreviewRow(row_number=index, data=data, errors=errors, warnings=warnings))

    _add_duplicate_isbn_issues(db, rows)
    invalid_rows = sum(1 for row in rows if row.errors)
    return ImportPreviewResponse(
        total_rows=len(rows),
        valid_rows=len(rows) - invalid_rows,
        invalid_rows=invalid_rows,
        rows=rows,
        errors=[issue for row in rows for issue in row.errors],
    )


def confirm_import(
    db: Session,
    content: bytes,
    file_format: str,
    field_mapping: dict[str, str] | None = None,
    current_user_id: int | None = None,
) -> ImportConfirmResponse:
    preview = preview_import(db, content, file_format, field_mapping)
    if preview.errors:
        raise ApiError(
            "VALIDATION_ERROR",
            "导入文件存在错误，未写入任何数据",
            status_code=422,
            details=[issue.model_dump() for issue in preview.errors],
        )

    original_commit = db.commit
    try:
        db.commit = db.flush  # type: ignore[method-assign]
        for row in preview.rows:
            payload = BookCreate(**row.data)
            book_service.create_book(db, payload, current_user_id=current_user_id)
        db.commit = original_commit  # type: ignore[method-assign]
        original_commit()
    except Exception:
        db.commit = original_commit  # type: ignore[method-assign]
        db.rollback()
        raise

    return ImportConfirmResponse(imported_count=preview.valid_rows, skipped_count=0, errors=[])


def _book_to_export_row(book: Book) -> dict[str, Any]:
    return {
        "title": book.title,
        "subtitle": book.subtitle,
        "author": book.author,
        "translator": book.translator,
        "publisher": book.publisher,
        "publish_year": book.publish_year,
        "isbn": book.isbn,
        "original_isbn": book.original_isbn,
        "language": book.language,
        "pages": book.pages,
        "price_cents": book.price_cents,
        "binding": book.binding,
        "series": book.series,
        "cover_url": book.cover_url,
        "summary": book.summary,
        "author_intro": book.author_intro,
        "category_id": book.category_id,
        "category_code": book.category.code if book.category else None,
        "category_name": book.category.name if book.category else None,
        "location_id": book.location_id,
        "location_full_path": book.location.full_path if book.location else None,
        "status": book.status,
        "read_status": book.read_status,
        "rating": book.rating,
        "is_favorite": book.is_favorite,
        "tag_names": ";".join(book_tag.tag.name for book_tag in book.book_tags),
        "note": book.note,
    }


def list_export_rows(db: Session) -> list[dict[str, Any]]:
    books = (
        db.query(Book)
        .options(
            joinedload(Book.category),
            joinedload(Book.location),
            selectinload(Book.book_tags).joinedload(BookTag.tag),
        )
        .order_by(Book.id.asc())
        .all()
    )
    return [_book_to_export_row(book) for book in books]


def export_csv(rows: Iterable[dict[str, Any]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=EXPORT_FIELDS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8-sig")


def export_json(rows: Iterable[dict[str, Any]]) -> bytes:
    return json.dumps({"items": list(rows)}, ensure_ascii=False, indent=2).encode("utf-8")


def export_xlsx(rows: Iterable[dict[str, Any]]) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ApiError("VALIDATION_ERROR", "Excel 导出需要安装 openpyxl 依赖", status_code=422) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "books"
    sheet.append(EXPORT_FIELDS)
    for row in rows:
        sheet.append([row.get(field) for field in EXPORT_FIELDS])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_books(db: Session, file_format: str) -> tuple[bytes, str, str]:
    rows = list_export_rows(db)
    normalized = file_format.lower()
    if normalized == "csv":
        return export_csv(rows), "text/csv; charset=utf-8", "books.csv"
    if normalized == "json":
        return export_json(rows), "application/json", "books.json"
    if normalized in {"xlsx", "excel"}:
        return (
            export_xlsx(rows),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "books.xlsx",
        )
    raise ApiError("VALIDATION_ERROR", "format 仅支持 csv、json、xlsx", status_code=422)
